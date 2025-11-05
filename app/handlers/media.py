"""Обработчики медиа и текста."""
import asyncio
import json
import uuid
from typing import Dict, List, Set, Optional
from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.types import Message
from app.utils.logging import get_logger
from app.services.tg_files import TelegramFileService
from app.services.webhook_client import WebhookClient
from app.services.prefs import PreferencesService
from app.models.payload import WebhookPayload, Creative, ChatInfo, UserInfo, MessageInfo, BatchInfo
from app.models.payload import TextsPayload
from app.utils.env import config
import httpx
from io import BytesIO

logger = get_logger(__name__)
router = Router()

# Глобальные переменные для управления burst-режимом
user_buffers: Dict[int, List[Message]] = {}
user_timers: Dict[int, asyncio.Task] = {}
media_groups: Dict[str, List[Message]] = {}
media_group_timers: Dict[str, asyncio.Task] = {}

# Сервисы
tg_files_service = TelegramFileService(None)  # Будет инициализирован в main
webhook_client = WebhookClient()
prefs_service = PreferencesService()

# Инициализация будет выполнена в main.py

@router.message(F.media_group_id & F.photo)
async def handle_media_group(message: Message):
    """Обработчик альбомов фото (media groups)."""
    media_group_id = message.media_group_id
    user_id = message.from_user.id
    
    # Добавляем сообщение в группу
    if media_group_id not in media_groups:
        media_groups[media_group_id] = []
    media_groups[media_group_id].append(message)
    
    # Отменяем предыдущий таймер для этой группы
    if media_group_id in media_group_timers:
        media_group_timers[media_group_id].cancel()
    
    # Создаем новый таймер
    timer_task = asyncio.create_task(
        process_media_group_after_delay(media_group_id, user_id)
    )
    media_group_timers[media_group_id] = timer_task
    
    logger.info(f"📦 Добавлено фото в media group {media_group_id}")

async def process_media_group_after_delay(media_group_id: str, user_id: int):
    """Обработать media group после задержки."""
    await asyncio.sleep(1.5)  # Ждем 1.5 секунды для сбора всех сообщений
    
    if media_group_id in media_groups:
        messages = media_groups[media_group_id]
        del media_groups[media_group_id]
        
        if media_group_id in media_group_timers:
            del media_group_timers[media_group_id]
        
        await process_messages_batch(messages, user_id, "media_group")
        
        logger.info(f"📦 Обработан media group {media_group_id} с {len(messages)} сообщениями")

@router.message(F.photo & ~F.media_group_id)
async def handle_single_photo(message: Message):
    """Обработчик одиночных фото (не в media group)."""
    user_id = message.from_user.id
    
    # Добавляем в burst-буфер
    if user_id not in user_buffers:
        user_buffers[user_id] = []
    
    user_buffers[user_id].append(message)
    
    # Отменяем предыдущий таймер для пользователя
    if user_id in user_timers:
        user_timers[user_id].cancel()
    
    # Создаем новый таймер
    timer_task = asyncio.create_task(
        process_user_buffer_after_delay(user_id)
    )
    user_timers[user_id] = timer_task
    
    logger.info(f"📎 Добавлено одиночное фото в буфер пользователя {user_id}")

@router.message(F.text)
async def handle_texts(message: Message):
    """Обработчик текстовых объявлений.
    - Разбивает по строкам
    - Отправляет массив строк на текстовый вебхук выбранного сервиса
    """
    user_id = message.from_user.id
    service = prefs_service.get_user_service(user_id)
    webhook_url = config.get_text_webhook_url(service)
    if not webhook_url:
        await message.answer("❌ Текстовый вебхук не настроен для выбранного сервиса")
        return
    # Пингуем для логов, но не блокируем отправку (prod может возвращать 401/405)
    try:
        _ = await webhook_client.send_ping(webhook_url)
    except Exception:
        pass
    lines = [line.strip() for line in (message.text or "").splitlines()]
    texts = [line for line in lines if line]
    if not texts:
        await message.answer("⚠️ Текст не найден для отправки")
        return
    idem = webhook_client.generate_idempotency_key(str(message.message_id), 1)
    placement = prefs_service.get_user_placement(user_id)
    chat = {
        "chat_id": message.chat.id,
        "type": message.chat.type,
        "title": message.chat.title,
    }
    from_ = {
        "user_id": message.from_user.id,
        "username": message.from_user.username,
    }
    ok = await webhook_client.send_texts(
        texts=texts,
        webhook_url=webhook_url,
        service=service,
        chat=chat,
        from_=from_,
        placement=placement,
        idempotency_key=idem,
    )
    if ok:
        await message.answer(f"✅ Отправлено {len(texts)} текстов на {service.title()}")
    else:
        await message.answer("❌ Не удалось отправить тексты")

@router.message(F.document)
async def handle_excel(message: Message):
    """Обработчик Excel-файлов (.xlsx):
    - Скачивает файл
    - Собирает все ячейки, кроме первой строки (заголовка), по всем листам
    - Отправляет массив непустых строк на текстовый вебхук выбранного сервиса
    """
    if not message.document:
        return
    filename = message.document.file_name or ""
    mime = message.document.mime_type or ""
    is_xlsx = filename.lower().endswith(".xlsx") or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if not is_xlsx:
        return  # игнорируем прочие документы
    user_id = message.from_user.id
    service = prefs_service.get_user_service(user_id)
    webhook_url = config.get_text_webhook_url(service)
    if not webhook_url:
        await message.answer("❌ Текстовый вебхук не настроен для выбранного сервиса")
        return
    # получаем URL файла и скачиваем
    file_url = await tg_files_service.get_file_url(message.document.file_id)
    if not file_url:
        await message.answer("❌ Не удалось получить файл")
        return
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(config.HTTP_TIMEOUT_SECONDS)) as client:
            resp = await client.get(file_url)
            resp.raise_for_status()
            data = BytesIO(resp.content)
    except Exception as e:
        logger.error(f"❌ Ошибка скачивания Excel: {e}")
        await message.answer("❌ Ошибка скачивания файла")
        return
    # парсим xlsx (ленивый импорт openpyxl)
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(filename=data, read_only=True, data_only=True)
        texts: List[str] = []
        for ws in wb.worksheets:
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue  # пропускаем заголовок
                for cell in row:
                    if cell is None:
                        continue
                    s = str(cell).strip()
                    if s:
                        texts.append(s)
    except Exception as e:
        logger.error(f"❌ Ошибка обработки Excel: {e}")
        await message.answer("❌ Ошибка обработки Excel")
        return
    if not texts:
        await message.answer("⚠️ Не найден текст в Excel")
        return
    # Пингуем для логов, но не блокируем отправку
    try:
        _ = await webhook_client.send_ping(webhook_url)
    except Exception:
        pass
    idem = webhook_client.generate_idempotency_key(str(message.message_id), 1)
    placement = prefs_service.get_user_placement(user_id)
    chat = {
        "chat_id": message.chat.id,
        "type": message.chat.type,
        "title": message.chat.title,
    }
    from_ = {
        "user_id": message.from_user.id,
        "username": message.from_user.username,
    }
    ok = await webhook_client.send_texts(
        texts=texts,
        webhook_url=webhook_url,
        service=service,
        chat=chat,
        from_=from_,
        placement=placement,
        idempotency_key=idem,
    )
    if ok:
        await message.answer(f"✅ Отправлено {len(texts)} текстов из Excel на {service.title()}")
    else:
        await message.answer("❌ Не удалось отправить тексты из Excel")

async def process_user_buffer_after_delay(user_id: int):
    """Обработать буфер пользователя после задержки."""
    # Ждем debounce время
    await asyncio.sleep(config.BURST_DEBOUNCE_SECS)
    
    if user_id in user_buffers and user_buffers[user_id]:
        messages = user_buffers[user_id].copy()
        user_buffers[user_id] = []
        
        if user_id in user_timers:
            del user_timers[user_id]
        
        await process_messages_batch(messages, user_id, "debounce")
        
        logger.info(f"⚡ Обработан burst пользователя {user_id} с {len(messages)} сообщениями")

async def process_messages_batch(messages: List[Message], user_id: int, grouping: str):
    """Обработать пакет сообщений."""
    if not messages:
        return
    
    # Получаем сервис пользователя
    service = prefs_service.get_user_service(user_id)
    placement = prefs_service.get_user_placement(user_id)
    webhook_url = config.get_webhook_url(service)
    
    if not webhook_url:
        await messages[0].answer("❌ Не удалось определить URL вебхука")
        return
    
    # Извлекаем креативы
    creatives = await tg_files_service.extract_creatives_from_messages(messages)
    
    if not creatives:
        await messages[0].answer("❌ Не удалось извлечь креативы из сообщений")
        return
    
    # Собираем download URLs
    download_urls = [creative.download_url for creative in creatives if creative.download_url]
    
    # Создаем batch info
    batch_id = str(uuid.uuid4())
    
    # Разбиваем на чанки если нужно
    max_per_batch = config.MAX_CREATIVES_PER_BATCH
    chunks = [creatives[i:i + max_per_batch] for i in range(0, len(creatives), max_per_batch)]
    
    success_count = 0
    
    for seq, chunk in enumerate(chunks, 1):
        # Создаем payload
        payload = create_webhook_payload(
            messages=messages,
            creatives=chunk,
            download_urls=download_urls,
            service=service,
            batch_id=batch_id,
            seq=seq,
            total=len(chunks),
            grouping=grouping,
            placement=placement
        )
        
        # Отправляем на вебхук
        idempotency_key = webhook_client.generate_idempotency_key(batch_id, seq)
        success = await webhook_client.send_payload(payload, webhook_url, idempotency_key)
        
        if success:
            success_count += 1
            # Сохраняем payload для retry
            prefs_service.save_last_payload(user_id, payload.model_dump_json())
    
    # Уведомляем пользователя
    if success_count == len(chunks):
        await messages[0].answer(f"✅ Отправлено {len(creatives)} креативов на {service.title()}")
    else:
        await messages[0].answer(f"⚠️ Отправлено {success_count}/{len(chunks)} пакетов на {service.title()}")

def create_webhook_payload(
    messages: List[Message],
    creatives: List[Creative],
    download_urls: List[str],
    service: str,
    batch_id: str,
    seq: int,
    total: int,
    grouping: str,
    placement: Optional[str] = None
) -> WebhookPayload:
    """Создать payload для вебхука."""
    first_message = messages[0]
    
    # Информация о чате
    chat_info = ChatInfo(
        chat_id=first_message.chat.id,
        type=first_message.chat.type,
        title=first_message.chat.title
    )
    
    # Информация о пользователе
    user_info = UserInfo(
        user_id=first_message.from_user.id,
        username=first_message.from_user.username
    )
    
    # Информация о сообщении
    message_info = MessageInfo(
        message_id=first_message.message_id,
        date_ts=int(first_message.date.timestamp()),
        media_group_id=first_message.media_group_id
    )
    
    # ID всех сообщений
    message_ids = [msg.message_id for msg in messages]
    
    # Информация о батче
    batch_info = BatchInfo(
        batch_id=batch_id,
        seq=seq,
        total=total,
        grouping=grouping
    )
    
    return WebhookPayload(
        service=service,
        chat=chat_info,
        from_=user_info,
        message=message_info,
        message_ids=message_ids,
        creatives=creatives,
        download_urls=download_urls,
        batch=batch_info,
        placement=placement
    )
